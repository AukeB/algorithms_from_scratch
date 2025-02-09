import os
import logging
from pathlib import Path
from openpyxl import load_workbook
from PIL import Image
from collections import namedtuple

Size = namedtuple("Size", ["width", "height"])


class BitmapUtils:
    """ """

    def __init__(self, config: dict, file_name: str) -> None:
        """ """
        self.config = config
        self.file_name = file_name

    def _hex_to_rgb(self, color_hex):
        """ """
        r = int(color_hex[0:2], 16)
        g = int(color_hex[2:4], 16)
        b = int(color_hex[4:6], 16)

        return (r, g, b)

    def read_bitmap_from_excel(
        self,
    ) -> tuple[list[list[str]], list[list[str]]]:
        """ """
        relative_dir_path: str = self.config["paths"]["bitmaps_dir"]
        relative_file_path = os.path.join(relative_dir_path, self.file_name)

        # This assumes the bitmap is always on the first tab of the excel file.
        sheet = load_workbook(relative_file_path).worksheets[0]
        size_width: int = self.config["bitmaps"][Path(self.file_name).stem]["dimensions"][0]
        size_height: int = self.config["bitmaps"][Path(self.file_name).stem]["dimensions"][1]
        bitmap = []

        for row in sheet.iter_rows(
            min_row=1,
            max_row=size_width,
            min_col=1,
            max_col=size_height,
        ):
            bitmap_row = []

            for cell in row:
                color_hex = cell.fill.start_color.index
                color_rgb = self._hex_to_rgb(
                    color_hex=color_hex[2:]
                )  # First two elements contain transparency/alpha/opacity.
                bitmap_row.append(color_rgb)

            bitmap.append(bitmap_row)
        if self.config["png_bitmap"]["export"]:
            self.export_bitmap_as_png(bitmap=bitmap, file_name=self.file_name)

        return bitmap

    def export_bitmap_as_png(
        self,
        bitmap: list[list[tuple[int, int, int]]],
        file_name: str,
        output_file_extension: str = ".png",
    ) -> None:
        """ """
        bitmap_cell_size: int = self.config["png_bitmap"]["cell_size"]
        bitmap_dimensions = Size(len(bitmap), len(bitmap[0]))
        img_width = bitmap_cell_size * bitmap_dimensions.width
        img_height = bitmap_cell_size * bitmap_dimensions.height

        img = Image.new(
            "RGB",
            (img_width, img_height),
            color=tuple(self.config["png_bitmap"]["default_background_color"]),
        )

        for y, row in enumerate(bitmap):
            for x, color in enumerate(row):
                for dx in range(bitmap_cell_size):
                    for dy in range(bitmap_cell_size):
                        img.putpixel((x * bitmap_cell_size + dx, y * bitmap_cell_size + dy), color)

        image_filename: str = (
            self.config["paths"]["bitmaps_dir"] + Path(file_name).stem + output_file_extension
        )
        img.save(image_filename)

        logging.info(f"Bitmap exported as {image_filename}")

    def create_color_mapping(self, bitmap):
        """ """
        color_mapping = {}
        current_char = "A"

        for row in bitmap:
            for color in row:
                if color not in color_mapping:
                    color_mapping[color] = current_char
                    # Fails when the maximum number of codepoints within Unicode has been
                    # reached (somewhere at 1.1M), so that's fine.
                    current_char = chr(ord(current_char) + 1)

        return color_mapping

    def apply_color_mapping(self, bitmap, color_mapping):
        """ """
        return [[color_mapping[color] for color in row] for row in bitmap]
